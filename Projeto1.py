import sqlite3
from openpyxl import load_workbook

class baseDados:
    def __init__(self, arquivo):
        self.connec = sqlite3.connect(arquivo)
        self.cursor = self.connec.cursor()
        self.listaItens = []

    def inserir(self, cod, nome, preco, qtd = None, val = None):
        consulta = 'INSERT OR IGNORE INTO itens (cod, nome, preco, qtd, val) VALUES (?, ?, ?, ?, ?)'
        self.cursor.execute(consulta, (cod, nome, preco, qtd, val))
        self.connec.commit()

    def getItens(self,planilhaNova = False):
        wb = load_workbook(filename='tabela de preços 2021.xlsx', data_only=True)
        ws = wb.active
        row_count = int(ws.max_row)
        for c in range(1, row_count + 1):
            numA = "A" + str(c)
            numB = "B" + str(c)
            numC = "C" + str(c)
            cellA = ws[numA].value
            cellB = ws[numB].value
            cellC = ws[numC].value
            item = []
            if isinstance(cellA, int):
                item.append(cellA)
                item.append(cellB)
                item.append(float(cellC))
                self.listaItens.append(item)
                if planilhaNova:
                    baseDados.inserir(self,cellA,cellB,cellC)
        for c in range(1, row_count + 1):
            numJ = "J" + str(c)
            numK = "K" + str(c)
            numL = "L" + str(c)
            cellJ = ws[numJ].value
            cellK = ws[numK].value
            cellL = ws[numL].value
            item = []
            if isinstance(cellJ, int):
                item.append(cellJ)
                item.append(cellK)
                item.append(float(cellL))
                self.listaItens.append(item)
                if planilhaNova:
                    baseDados.inserir(self,cellJ,cellK,cellL)

    def attBase(self):
        self.getItens()
        listaCod = []
        listaPrc = []
        consulta = 'SELECT * FROM itens'
        self.cursor.execute(consulta)
        for l in self.cursor.fetchall():
            codigo, nome, preco, qtd, validade = l
            listaCod.append(codigo)
            listaPrc.append(preco)
        for item in self.listaItens:
            itemCod, itemNome, precoNovo = item
            if itemCod in listaCod:
                indice = listaCod.index(itemCod)
                precoAntigo = listaPrc[indice]
                if str(precoAntigo) != str(precoNovo):
                    update = 'UPDATE itens SET preco = ? WHERE cod = ?'
                    self.cursor.execute(update, (precoNovo, itemCod))
                    self.connec.commit()
                    print(f'Preço do item "{itemCod}" atualizado de {precoAntigo} para {precoNovo}!')
            else:
                self.inserir(itemCod, itemNome, precoNovo)
                print(f'Item {itemCod}: {itemNome} inserido na base de dados')

    def listar(self):
        self.cursor.execute('SELECT * FROM itens')
        for l in self.cursor.fetchall():
            print(l)

    def editVal(self,cod, qtd, val, op = ('+','-')):
        item = baseDados.getItemfromId(self, cod)
        print(item)
        if baseDados.verificar(self,cod,val) != None:
            qtdItem = item[3]
            if qtdItem == None:
                qtdItem = 0
            if op == '+':
                quantidade = qtdItem + int(qtd)
            elif op == '-':
                quantidade = qtdItem - int(qtd)
            print(baseDados.verificar(self, cod, val))
            print(f'Validade "{val}" já cadastrada, quantidade atualizada para {quantidade}')
            if quantidade == 0:
                update = 'UPDATE itens SET qtd = ?, val = ? WHERE cod = ? AND val = ?'
                self.cursor.execute(update, (None, None, cod, val))
                self.connec.commit()
            else:
                update = 'UPDATE itens SET qtd = ? WHERE cod = ? AND val = ?'
                self.cursor.execute(update, (quantidade, cod, val))
                self.connec.commit()
        else:
            baseDados.inserir(self, cod, item[1], item[2], qtd, val)
            print(f'Data {val} inserida para o item "{cod}"')

    def qtdTotal(self,cod):
        item = []
        soma = 0
        consulta = 'SELECT * FROM itens'
        self.cursor.execute(consulta)
        for l in self.cursor.fetchall():
            codigo, nome, preco, qtd, validade = l
            if codigo == cod:
                item.append(l)
        for i in item:
            if i[3] == None:
                soma += 0
            else:
                soma += i[3]
        return soma

    def verificar(self,cod,val):
        item = None
        consulta = 'SELECT * FROM itens'
        self.cursor.execute(consulta)
        for l in self.cursor.fetchall():
            codigo, nome, preco, qtd, validade = l
            if codigo == cod and validade == val:
                item = l
        return item

    def getItemfromId(self, cod):
        item = None
        consulta = 'SELECT * FROM itens'
        self.cursor.execute(consulta)
        for l in self.cursor.fetchall():
            codigo, nome, preco, qtd, validade = l
            if codigo == cod:
                item = l
        return item

    def fechar(self):
        self.connec.close()
        self.cursor.close()



itens = baseDados('itensBaseDados.db')
# itens.attBase()
# print(itens.qtdTotal(628))
itens.editVal(628,2,'26/02/22','+')